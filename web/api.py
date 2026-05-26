"""FastAPI backend for the external React frontend.

Mirrors the Streamlit operator UI's actions but exposed as JSON over
HTTPS so the frontend (or any authenticated client) can drive the
pipeline from a browser. Every mutating action shells out to the
matching `scripts/*.py` so the workspace lock + audit + backup story
is unchanged.

Auth (resolved per request in `web/deps.py::require_auth`):
  1. Supabase HS256 JWT signed with `SUPABASE_JWT_SECRET`, sent as
     `Authorization: Bearer <jwt>`. Audience `authenticated`. The
     verified claims surface as a Principal (user_id, email, role).
  2. During the cutover window: the legacy shared `API_KEY` is also
     accepted when `AUTH_ALLOW_API_KEY_FALLBACK=true` AND
     `API_KEY_FALLBACK_USER_ID` is set (so the shared key gets
     attributed to a real tenant rather than silently
     mis-attributing traffic). Removed once the frontend stops
     sending the legacy key.

Multi-tenant routing: when `WORKSPACE_PER_USER=true` is set, every
request is scoped to the principal's workspace under
`${WORKSPACES_ROOT}/{user_id}/` (provisioned from a template on
first use). Legacy single-workspace deployments keep working by
leaving the flag unset and pinning `INVESTOR_WORKSPACE` directly.

CORS allow-list lives in `CORS_ORIGINS` (comma-separated) plus
optional `CORS_ORIGIN_REGEX` for ephemeral preview origins (e.g.
Lovable's `*--<project-id>.lovableproject.com`). The `Authorization`
header is allow-listed so JWTs flow through unchanged.

OpenAPI spec is auto-generated at /openapi.json (and used by the
`web/openapi.json` dump script in CI for the frontend's type
generator).
"""
from __future__ import annotations

import hmac
import os
import pathlib
import re
import subprocess
import sys
from typing import Any, Literal

from fastapi import (
    Depends, FastAPI, File, Form, HTTPException, Header,
    Query, Request, UploadFile,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from pydantic import BaseModel, Field
from sqlalchemy import desc, select

REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from core.approval.gate import (  # noqa: E402
    ApprovalGate,
    can_approve_draft,
    split_blockers,
)
from core.approval.persistence import (  # noqa: E402
    REVIEWABLE_STATES,
    approved_for_send,
    pending_review,
)
from core import gmail_oauth  # noqa: E402
from core.config_loader import load_workspace  # noqa: E402
from core.db import (  # noqa: E402
    crm_connections,
    draft_snoozes,
    email_drafts,
    get_engine,
    outreach_events,
    partner_pipeline,
    partner_score_summaries,
    partners,
    runs,
    today_picks,
    workspace_settings,
)


# ---------- pydantic response schemas ----------
#
# BlockerInfo / GateInfo / DraftView / CommandResult moved to
# web/deps.py (refactor #16) so the per-feature routers under
# web/routers/ can return them without a web.api import cycle.
# Re-exported here so existing call sites in this file keep
# working with no churn.
from web.deps import (  # noqa: E402
    BlockerInfo, CommandResult, DraftView, GateInfo,
    rationale_by_partner as _rationale_by_partner,
)


class ApproveBody(BaseModel):
    notes: str = Field(min_length=1, description="Operator rationale; required for audit.")
    override_blockers: bool = False


class RejectBody(BaseModel):
    notes: str = Field(min_length=1)


class SetEmailBody(BaseModel):
    email: str = Field(min_length=3)


class CheckReadyResult(BaseModel):
    phase: str
    stdout: str
    blocked: bool
    return_code: int


# ---------- onboarding wizard schemas ----------

class ConfigInfo(BaseModel):
    """Snapshot the onboarding wizard polls."""
    mode: Literal["fixture", "dry_run", "production"]
    gmail_connected: bool
    # Build Session 13: the same OAuth token now requests Drive scope
    # so meeting-prep briefs can be auto-pushed to the operator's
    # Drive. drive_connected is True only when the saved token
    # actually carries the drive.file scope -- legacy gmail-only
    # tokens read as gmail_connected=True, drive_connected=False, and
    # the wizard prompts a re-consent.
    drive_connected: bool
    # google_connected = both scopes present. Equivalent to (gmail &&
    # drive) but exposed as its own field so the wizard's "Connect
    # Google" button can render a single boolean state.
    google_connected: bool
    # True when company.name + company.one_liner are both non-empty,
    # i.e. the operator finished Step 1 of the wizard.
    company_configured: bool


# NOTE: GoogleStatus / GmailStatus / GmailConnectResponse moved to
# web/routers/google.py (Build Session 17). They are not imported
# back here because nothing else in this module references them.


class CompanyProfile(BaseModel):
    """Step 1 onboarding form. Flat shape (the UI's contract).
    Only `name` and `one_liner` are required; everything else
    defaults to "" / None / [] so a half-filled form still round-trips
    cleanly through PUT -> GET.
    """
    # Identity.
    name: str = ""
    one_liner: str = ""
    website: str = ""
    founded_year: int | None = None
    hq_location: str = ""
    # Pitch.
    stage: str = ""
    sectors: list[str] = Field(default_factory=list)
    business_model: str = ""
    problem: str = ""
    solution: str = ""
    differentiators: str = ""
    why_now: str = ""
    traction: str = ""
    # Round.
    round_amount_usd: int | None = None
    round_instrument: str = ""
    round_valuation_usd: int | None = None
    round_close_target: str = ""
    # Investor fit.
    target_check_min_usd: int | None = None
    target_check_max_usd: int | None = None
    target_stages: list[str] = Field(default_factory=list)
    target_sectors: list[str] = Field(default_factory=list)
    target_geographies: list[str] = Field(default_factory=list)
    desired_traits: list[str] = Field(default_factory=list)
    # Anti-criteria.
    excluded_sectors: list[str] = Field(default_factory=list)
    excluded_geographies: list[str] = Field(default_factory=list)
    do_not_contact: list[str] = Field(default_factory=list)
    # Voice + outreach.
    founder_name: str = ""
    founder_title: str = ""
    founder_email: str = ""
    signature: str = ""
    tone: str = ""
    scheduling_link: str = ""


class SetModeBody(BaseModel):
    mode: Literal["fixture", "dry_run", "production"]


class RunRow(BaseModel):
    run_id: int
    stage: str | None
    started_at: str | None
    completed_at: str | None
    processed: int | None
    succeeded: int | None
    failed: int | None
    skipped: int | None
    error_summary: str | None


# ---------- helpers ----------
#
# Build Session 17: the shared helpers (require_auth, _engine_and_ws,
# _ws_path, _api_key, _actor, _allow_example_domains_args, _run_cli)
# moved to web/deps.py so per-feature routers under web/routers/
# can import them without creating a circular import back to this
# module. The names are re-exported below so existing callers in
# this file keep working unchanged.

from web.deps import (  # noqa: E402
    _actor,
    _allow_example_domains_args,
    _api_key,
    _engine_and_ws,
    _run_cli,
    _ws_path,
    current_principal,
    require_auth,
)


# _gate_to_dict + _serialize_draft moved to web/deps.py as
# gate_to_dict + serialize_draft (refactor #16) so coach
# router can reuse them. Re-export under the old leading-
# underscore names so existing call sites in this file
# keep working.
from web.deps import gate_to_dict as _gate_to_dict  # noqa: E402
from web.deps import serialize_draft as _serialize_draft  # noqa: E402


# Coach B1 send-pace + review #11 discovery-opt-in helpers moved
# to web/routers/coach.py (refactor #16). Re-export the ones
# /pipeline/sources still calls (the dual-write opt-in check)
# so the call site doesn't need to update.
from web.routers.coach import (  # noqa: E402
    _read_discovery_opt_in,
)


# ---------- app ----------

app = FastAPI(
    title="Investor Outreach API",
    version="1.0.0",
    description=(
        "Backend for the external React frontend "
        "(`awesome_investor_tool`). All mutating endpoints shell "
        "out to the matching scripts/*.py so the existing workspace "
        "lock + audit + backup story is preserved."
    ),
)

# CORS allow-list. Two env vars, used together:
#
#   CORS_ORIGINS         comma-separated exact origins.
#   CORS_ORIGIN_REGEX    optional regex matched against the Origin
#                        header. Use when the frontend host generates
#                        ephemeral preview URLs (e.g. Lovable spawns
#                        a new `*--<project-id>.lovableproject.com`
#                        per session). Browser-sent origin = scheme
#                        + host only -- regex must NOT include path.
#
# starlette OR's the two: a request matches if its origin is in the
# explicit list OR matches the regex.
#
# Wildcard fallback ("*") fires ONLY when neither env var is set --
# typical local-dev shape. Production must pin one or both; a regex
# combined with an empty explicit list is the right pattern for a
# host like Lovable.
_origins_raw = os.environ.get("CORS_ORIGINS", "")
_origins = [o.strip() for o in _origins_raw.split(",") if o.strip()]
_origin_regex = os.environ.get("CORS_ORIGIN_REGEX") or None
if not _origins and not _origin_regex:
    _origins = ["*"]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_origin_regex=_origin_regex,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
)


# Per-user routing middleware (review items #1 + #2):
#
# FastAPI runs sync dependencies in a threadpool, and contextvar
# mutations made there don't propagate back to the parent async
# task -- so setting `_CURRENT_USER_ID_VAR` from inside
# `current_principal` / `require_auth` (which are sync) doesn't
# reach the endpoint. The fix is to stamp the contextvar from
# an async middleware that runs in the parent task. Every
# downstream dependency + endpoint then inherits the value via
# the normal context-propagation rules.
#
# The middleware is best-effort -- it doesn't reject anything;
# `require_auth` still does the actual gating. We just pre-resolve
# the principal so the user_id is in the contextvar by the time
# `_engine_and_ws()` / `_ws_path()` look for it.
@app.middleware("http")
async def _stamp_user_id_contextvar(request: Request, call_next):
    from web.deps import (
        _CURRENT_USER_ID_VAR,
        _is_api_key_fallback_enabled,
        _jwt_secret,
        _principal_from_claims,
        _verify_supabase_jwt,
    )

    auth = request.headers.get("authorization") or ""
    if auth.lower().startswith("bearer "):
        token = auth.split(" ", 1)[1].strip()
        user_id: str | None = None
        if _jwt_secret() is not None:
            claims = _verify_supabase_jwt(token)
            if claims is not None:
                principal = _principal_from_claims(claims)
                if principal and principal.get("user_id"):
                    user_id = principal["user_id"]
        if user_id is None:
            # Legacy API_KEY -> the bound fallback user_id (if any).
            fallback_on = (
                _jwt_secret() is None or _is_api_key_fallback_enabled()
            )
            if fallback_on:
                expected = os.environ.get("API_KEY") or ""
                if expected and hmac.compare_digest(token, expected):
                    bound = os.environ.get("API_KEY_FALLBACK_USER_ID") or ""
                    if bound:
                        user_id = bound
        if user_id:
            _CURRENT_USER_ID_VAR.set(user_id)
    return await call_next(request)


# ---------- routers ----------
#
# Per-feature routers live under web/routers/. Build Session 17
# started the migration with the Google OAuth surface (4 endpoints);
# future PRs can extract the onboarding-wizard, review-queue, and
# pipeline-runner clusters into their own routers using the same
# pattern. Endpoints not yet extracted remain inline below.
from web.routers.google import router as google_router  # noqa: E402
from web.routers.admin import router as admin_router  # noqa: E402
from web.routers.hooks import router as hooks_router  # noqa: E402
from web.routers.crm import router as crm_router  # noqa: E402
from web.routers.coach import router as coach_router  # noqa: E402
from web.routers.investors import router as investors_router  # noqa: E402
from web.routers.cadence import router as cadence_router  # noqa: E402
from web.routers.sequences import router as sequences_router  # noqa: E402

app.include_router(google_router)
app.include_router(admin_router)
app.include_router(hooks_router)
app.include_router(crm_router)
app.include_router(coach_router)
app.include_router(investors_router)
app.include_router(cadence_router)
app.include_router(sequences_router)


@app.get("/", include_in_schema=False)
def root():
    """Public health check (no auth) -- useful for uptime monitors."""
    return {
        "service": "investor-outreach-api",
        "status": "ok",
        "version": app.version,
    }


# ---------- review / approve ----------

@app.get(
    "/review/pending",
    response_model=list[DraftView],
    summary="Drafts pending operator review",
    tags=["review"],
)
def get_pending(_auth: None = Depends(require_auth)) -> list[DraftView]:
    engine, ws = _engine_and_ws()
    drafts = pending_review(engine)
    with engine.begin() as conn:
        email_by_pid = {
            r.partner_id: r.email or ""
            for r in conn.execute(
                select(partners.c.partner_id, partners.c.email),
            )
        }
        rationale_by_pid = _rationale_by_partner(conn)
    out: list[DraftView] = []
    for d in drafts:
        gate = can_approve_draft(
            ws,
            engine,
            int(d.draft_id),
            allow_example_domains=bool(_allow_example_domains_args()),
        )
        out.append(_serialize_draft(
            d,
            partner_email=email_by_pid.get(d.partner_id),
            gate=_gate_to_dict(gate),
            rationale=rationale_by_pid.get(d.partner_id),
        ))
    return out


@app.get(
    "/drafts/approved",
    response_model=list[DraftView],
    summary="Drafts ready to send (approved live rows)",
    tags=["review"],
)
def get_approved(_auth: None = Depends(require_auth)) -> list[DraftView]:
    engine, _ = _engine_and_ws()
    drafts = approved_for_send(engine)
    with engine.begin() as conn:
        email_by_pid = {
            r.partner_id: r.email or ""
            for r in conn.execute(
                select(partners.c.partner_id, partners.c.email),
            )
        }
        rationale_by_pid = _rationale_by_partner(conn)
    return [
        _serialize_draft(
            d,
            partner_email=email_by_pid.get(d.partner_id),
            gate=None,  # gate is checked on approve; the queue is post-gate
            rationale=rationale_by_pid.get(d.partner_id),
        )
        for d in drafts
    ]


@app.post(
    "/drafts/{draft_id}/approve",
    response_model=CommandResult,
    summary="Approve a draft (shells out to approve_draft.py)",
    tags=["mutations"],
)
def approve_draft(
    draft_id: int, body: ApproveBody,
    _auth: None = Depends(require_auth),
) -> CommandResult:
    cli = [
        "approve_draft.py", "--workspace", _ws_path(),
        "--draft-id", str(draft_id),
        "--notes", body.notes,
    ]
    cli.extend(_allow_example_domains_args())
    if body.override_blockers:
        cli.append("--override-blockers")
    res = _run_cli(*cli)
    if res.returncode != 0:
        raise HTTPException(
            400,
            detail={
                "error": "approve refused",
                "stdout": res.stdout,
                "stderr": res.stderr,
            },
        )
    return CommandResult(ok=True, stdout=res.stdout, stderr=res.stderr)


@app.post(
    "/drafts/{draft_id}/reject",
    response_model=CommandResult,
    summary="Reject a draft (shells out to reject_draft.py)",
    tags=["mutations"],
)
def reject_draft(
    draft_id: int, body: RejectBody,
    _auth: None = Depends(require_auth),
) -> CommandResult:
    res = _run_cli(
        "reject_draft.py", "--workspace", _ws_path(),
        "--draft-id", str(draft_id), "--notes", body.notes,
    )
    if res.returncode != 0:
        raise HTTPException(
            400,
            detail={
                "error": "reject failed",
                "stdout": res.stdout,
                "stderr": res.stderr,
            },
        )
    return CommandResult(ok=True, stdout=res.stdout, stderr=res.stderr)


# B1 Today + B2 Sent + B3 Replies + B4 Pipeline/Snoozes + review
# #11 discovery-opt-in endpoints moved to web/routers/coach.py
# (refactor #16 final). Schemas (TodayPickView, SendPaceBody/View,
# DiscoveryOptInBody/View, PipelineBody/View, SnoozeBody/View,
# SentItem, ReplyItem) and helpers (_read_send_pace,
# _write_send_pace, _read_discovery_opt_in, _write_discovery_opt_in,
# _parse_future_iso) moved alongside them.


# ---------- partner mutations ----------

@app.post(
    "/partners/{partner_id}/email",
    response_model=CommandResult,
    summary="Set a partner's email (shells out to set_partner_email.py)",
    tags=["mutations"],
)
def set_partner_email(
    partner_id: str, body: SetEmailBody,
    _auth: None = Depends(require_auth),
) -> CommandResult:
    res = _run_cli(
        "set_partner_email.py", "--workspace", _ws_path(),
        "--partner-id", partner_id, "--email", body.email,
    )
    if res.returncode != 0:
        raise HTTPException(
            400,
            detail={
                "error": "set_partner_email failed",
                "stdout": res.stdout,
                "stderr": res.stderr,
            },
        )
    return CommandResult(ok=True, stdout=res.stdout, stderr=res.stderr)


# ---------- gates / status ----------

@app.get(
    "/check_ready",
    response_model=CheckReadyResult,
    summary="Pre-flight check (review|send|gmail|attio)",
    tags=["status"],
)
def check_ready(
    phase: str = Query(
        default="send",
        pattern="^(review|send|gmail|attio)$",
        description="Which workflow phase to gate on.",
    ),
    _auth: None = Depends(require_auth),
) -> CheckReadyResult:
    res = _run_cli(
        "check_ready.py", "--workspace", _ws_path(),
        "--for", phase, *_allow_example_domains_args(),
    )
    return CheckReadyResult(
        phase=phase,
        stdout=res.stdout,
        blocked="BLOCKED" in res.stdout,
        return_code=res.returncode,
    )


@app.get(
    "/runs",
    response_model=list[RunRow],
    summary="Recent run entries",
    tags=["status"],
)
def get_runs(
    limit: int = Query(50, ge=1, le=500),
    _auth: None = Depends(require_auth),
) -> list[RunRow]:
    engine, _ = _engine_and_ws()
    with engine.begin() as conn:
        rows = list(conn.execute(
            select(
                runs.c.run_id, runs.c.stage, runs.c.started_at,
                runs.c.completed_at,
                runs.c.records_processed, runs.c.records_succeeded,
                runs.c.records_failed, runs.c.records_skipped,
                runs.c.error_summary,
            ).order_by(desc(runs.c.run_id)).limit(limit)
        ))
    return [
        RunRow(
            run_id=int(r.run_id),
            stage=r.stage,
            started_at=str(r.started_at) if r.started_at else None,
            completed_at=str(r.completed_at) if r.completed_at else None,
            processed=r.records_processed,
            succeeded=r.records_succeeded,
            failed=r.records_failed,
            skipped=r.records_skipped,
            error_summary=r.error_summary,
        )
        for r in rows
    ]


# ---------- export ----------

@app.get(
    "/send_queue.csv",
    summary="Build + download the send_queue CSV",
    tags=["export"],
    responses={200: {"content": {"text/csv": {}}}},
)
def send_queue_csv(_auth: None = Depends(require_auth)) -> FileResponse:
    """Calls export_send_queue.py to materialize the CSV, then
    streams the file. Returns 400 if the export refuses (e.g. no
    approved drafts; stale approvals not skipped)."""
    res = _run_cli(
        "export_send_queue.py", "--workspace", _ws_path(),
        *_allow_example_domains_args(),
    )
    if res.returncode != 0:
        raise HTTPException(
            400,
            detail={
                "error": "export refused",
                "stdout": res.stdout,
                "stderr": res.stderr,
            },
        )
    _, ws = _engine_and_ws()
    csv_path = pathlib.Path(ws.path) / "exports" / "send_queue.csv"
    if not csv_path.exists():
        raise HTTPException(
            500, f"export reported success but file is missing: {csv_path}",
        )
    return FileResponse(
        path=str(csv_path),
        media_type="text/csv",
        filename="send_queue.csv",
    )


# ---------- onboarding wizard ----------
#
# These 6 endpoints back the React frontend's /onboarding route. The
# wizard walks an operator from "fresh checkout" to "drafts ready for
# review" by flipping company.yaml out of fixture mode, running stages
# 6 + 7, and linking Gmail -- four things the dashboard could not do
# without shelling out / editing files on the API host. Conventions
# match the existing routes: Bearer auth on every endpoint (except the
# OAuth callback, see below), subprocess failures surface as HTTP 400
# with detail={error, stdout, stderr, returncode}.

_MODE_LINE = re.compile(r"^(mode:\s*)(\S+)(.*)$", re.MULTILINE)
# A top-level YAML key at column 0: alnum/underscore identifier
# followed by `:`. Used to find the start + end of the `company:`
# block when rewriting it in place.
_TOPLEVEL_KEY = re.compile(r"^([A-Za-z_][A-Za-z0-9_]*)\s*:")


def _load_yaml_safely(path: pathlib.Path) -> dict:
    """Parse YAML from disk, tolerating missing / empty files."""
    if not path.exists():
        return {}
    try:
        import yaml  # type: ignore
    except ImportError as exc:  # pragma: no cover - PyYAML is a hard dep
        raise HTTPException(500, f"PyYAML missing: {exc}")
    text = path.read_text(encoding="utf-8")
    if not text.strip():
        return {}
    try:
        data = yaml.safe_load(text)
    except Exception as exc:  # noqa: BLE001 - propagate parse errors
        raise HTTPException(
            500, f"company.yaml is not valid YAML: {exc}",
        )
    return data or {}


def _read_company_block(yaml_path: pathlib.Path) -> "CompanyProfile":
    """Build a CompanyProfile from disk.

    Reads the flat fields directly from `company:`, falling back to
    the legacy nested keys (`target_check_size_usd.{min,max}`,
    `current_traction.headline_metric`,
    `meeting_ask.preferred_scheduling_link`) when the flat counterpart
    is empty. This means a workspace that's only ever been edited
    through the CLI (test_workspace) surfaces sensible values to the
    UI without first being re-written via PUT.
    """
    data = _load_yaml_safely(yaml_path)
    c = (data.get("company") or {}) if isinstance(data, dict) else {}

    def _str(k: str, default: str = "") -> str:
        v = c.get(k)
        return v if isinstance(v, str) else default

    def _list(k: str) -> list[str]:
        v = c.get(k)
        if isinstance(v, list):
            return [str(x) for x in v]
        return []

    def _int(k: str) -> int | None:
        v = c.get(k)
        if isinstance(v, bool):  # bool is an int subclass; reject it
            return None
        if isinstance(v, int):
            return v
        return None

    # Legacy nested fallbacks.
    nested_check = c.get("target_check_size_usd") or {}
    legacy_check_min = (
        nested_check.get("min") if isinstance(nested_check, dict) else None
    )
    legacy_check_max = (
        nested_check.get("max") if isinstance(nested_check, dict) else None
    )
    legacy_traction = ""
    ct = c.get("current_traction")
    if isinstance(ct, dict):
        legacy_traction = ct.get("headline_metric") or ""
    legacy_scheduling = ""
    ma = c.get("meeting_ask")
    if isinstance(ma, dict):
        legacy_scheduling = ma.get("preferred_scheduling_link") or ""

    return CompanyProfile(
        name=_str("name"),
        one_liner=_str("one_liner"),
        website=_str("website"),
        founded_year=_int("founded_year"),
        hq_location=_str("hq_location"),
        stage=_str("stage"),
        sectors=_list("sectors"),
        business_model=_str("business_model"),
        problem=_str("problem"),
        solution=_str("solution"),
        differentiators=_str("differentiators"),
        why_now=_str("why_now"),
        traction=_str("traction") or legacy_traction,
        round_amount_usd=_int("round_amount_usd"),
        round_instrument=_str("round_instrument"),
        round_valuation_usd=_int("round_valuation_usd"),
        round_close_target=_str("round_close_target"),
        target_check_min_usd=(
            _int("target_check_min_usd")
            or (int(legacy_check_min) if isinstance(legacy_check_min, int) else None)
        ),
        target_check_max_usd=(
            _int("target_check_max_usd")
            or (int(legacy_check_max) if isinstance(legacy_check_max, int) else None)
        ),
        target_stages=_list("target_stages"),
        target_sectors=_list("target_sectors"),
        target_geographies=_list("target_geographies"),
        desired_traits=_list("desired_traits"),
        excluded_sectors=_list("excluded_sectors"),
        excluded_geographies=_list("excluded_geographies"),
        do_not_contact=_list("do_not_contact"),
        founder_name=_str("founder_name"),
        founder_title=_str("founder_title"),
        founder_email=_str("founder_email"),
        signature=_str("signature"),
        tone=_str("tone"),
        scheduling_link=_str("scheduling_link") or legacy_scheduling,
    )


def _company_dict_for_yaml(profile: "CompanyProfile") -> dict:
    """Render a CompanyProfile as the dict to dump under `company:`.

    Stores every flat field the UI sent, AND mirrors the three legacy
    nested keys the existing pipeline code still reads
    (core/round_fit.py, core/email/prompt.py, core/email/draft_routing.py).
    The mirror keeps stages 6/7 working unchanged on a workspace whose
    `company:` block was last written by this endpoint.
    """
    out: dict = {
        "name": profile.name,
        "one_liner": profile.one_liner,
        "website": profile.website,
        "founded_year": profile.founded_year,
        "hq_location": profile.hq_location,
        "stage": profile.stage,
        "sectors": list(profile.sectors),
        "business_model": profile.business_model,
        "problem": profile.problem,
        "solution": profile.solution,
        "differentiators": profile.differentiators,
        "why_now": profile.why_now,
        "traction": profile.traction,
        "round_amount_usd": profile.round_amount_usd,
        "round_instrument": profile.round_instrument,
        "round_valuation_usd": profile.round_valuation_usd,
        "round_close_target": profile.round_close_target,
        "target_check_min_usd": profile.target_check_min_usd,
        "target_check_max_usd": profile.target_check_max_usd,
        "target_stages": list(profile.target_stages),
        "target_sectors": list(profile.target_sectors),
        "target_geographies": list(profile.target_geographies),
        "desired_traits": list(profile.desired_traits),
        "excluded_sectors": list(profile.excluded_sectors),
        "excluded_geographies": list(profile.excluded_geographies),
        "do_not_contact": list(profile.do_not_contact),
        "founder_name": profile.founder_name,
        "founder_title": profile.founder_title,
        "founder_email": profile.founder_email,
        "signature": profile.signature,
        "tone": profile.tone,
        "scheduling_link": profile.scheduling_link,
    }
    # Legacy mirrors for the existing pipeline code paths.
    if profile.target_check_min_usd is not None or profile.target_check_max_usd is not None:
        out["target_check_size_usd"] = {
            "min": profile.target_check_min_usd or 0,
            "max": profile.target_check_max_usd or 0,
        }
    if profile.traction:
        out["current_traction"] = {"headline_metric": profile.traction}
    if profile.scheduling_link:
        out["meeting_ask"] = {
            "preferred_scheduling_link": profile.scheduling_link,
        }
    return out


def _write_company_block(
    yaml_path: pathlib.Path, profile: "CompanyProfile",
) -> None:
    """Rewrite (or append) the top-level `company:` block in place.

    Algorithm: find the line range spanned by `company:` (start =
    line that begins with `company:`, end = first top-level key
    that follows), replace it with a fresh YAML dump of just that
    one block, leave every byte outside the range untouched. The
    result preserves: the `mode:` line, sibling blocks
    (`raise_context`, `founder_voice`, `round_fit`, ...), and every
    comment that doesn't live inside the `company:` block.

    Comments INSIDE the company block are dropped (PyYAML doesn't
    round-trip them; ruamel.yaml would, but it's not a runtime dep).
    The brief mentions this is acceptable -- the company block is
    operator-mutable through the UI and rarely hand-edited.
    """
    try:
        import yaml  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(500, f"PyYAML missing: {exc}")

    company_dict = _company_dict_for_yaml(profile)
    rendered = yaml.safe_dump(
        {"company": company_dict},
        sort_keys=False, allow_unicode=True, default_flow_style=False,
    )
    if not rendered.endswith("\n"):
        rendered += "\n"

    if not yaml_path.exists() or not yaml_path.read_text(encoding="utf-8").strip():
        yaml_path.parent.mkdir(parents=True, exist_ok=True)
        yaml_path.write_text(rendered, encoding="utf-8")
        return

    original = yaml_path.read_text(encoding="utf-8")
    lines = original.splitlines(keepends=True)
    start: int | None = None
    end: int | None = None
    for i, line in enumerate(lines):
        m = _TOPLEVEL_KEY.match(line)
        if not m:
            continue
        if start is None:
            if m.group(1) == "company":
                start = i
            continue
        # Already inside the company block; first top-level key after
        # start closes the range.
        end = i
        break
    if start is None:
        # Append at the end with a blank-line separator so the new
        # block is visually distinct from whatever came before.
        sep = "" if original.endswith("\n\n") or not original else "\n"
        yaml_path.write_text(original + sep + rendered, encoding="utf-8")
        return
    if end is None:
        end = len(lines)
    new_text = "".join(lines[:start]) + rendered + "".join(lines[end:])
    yaml_path.write_text(new_text, encoding="utf-8")


def _read_mode_from_yaml(yaml_path: pathlib.Path) -> str:
    """Wizard-facing mode value.

    Preserve the full on-disk vocabulary {fixture, dry_run, production}; the
    browser UI needs dry_run as a real safe pilot state.
    """
    if not yaml_path.exists():
        raise HTTPException(500, f"company.yaml missing at {yaml_path}")
    text = yaml_path.read_text(encoding="utf-8")
    m = _MODE_LINE.search(text)
    if not m:
        # config_loader defaults absence to dry_run.
        return "dry_run"
    value = m.group(2).strip().strip("\"'")
    if value in {"fixture", "dry_run", "production"}:
        return value
    return "dry_run"


def _write_mode_to_yaml(yaml_path: pathlib.Path, new_mode: str) -> None:
    """Replace (or insert) the top-level `mode:` line in company.yaml
    in place, preserving comments and surrounding formatting. A regex
    edit avoids re-rendering the whole document the way PyYAML's
    round-trip would (PyYAML drops comments; ruamel.yaml is not in our
    deps). Only one short line ever changes, so a string-level edit is
    sufficient."""
    if not yaml_path.exists():
        raise HTTPException(500, f"company.yaml missing at {yaml_path}")
    text = yaml_path.read_text(encoding="utf-8")
    if _MODE_LINE.search(text):
        new_text = _MODE_LINE.sub(
            lambda m: f"{m.group(1)}{new_mode}{m.group(3)}",
            text, count=1,
        )
    else:
        # Insert after the leading run of comment / blank lines so the
        # mode declaration sits above the actual config block.
        lines = text.splitlines()
        insert_at = 0
        for i, line in enumerate(lines):
            stripped = line.lstrip()
            if stripped and not stripped.startswith("#"):
                insert_at = i
                break
        new_line = f"mode: {new_mode}"
        # Pad with a blank line on either side so the inserted line
        # reads as its own block instead of glomming onto the next key.
        to_insert = [new_line, ""]
        if insert_at > 0 and lines[insert_at - 1].strip() != "":
            to_insert = ["", *to_insert]
        for offset, line in enumerate(to_insert):
            lines.insert(insert_at + offset, line)
        new_text = "\n".join(lines)
        if text.endswith("\n") and not new_text.endswith("\n"):
            new_text += "\n"
    yaml_path.write_text(new_text, encoding="utf-8")


@app.get(
    "/config",
    response_model=ConfigInfo,
    summary="Onboarding wizard config snapshot",
    tags=["onboarding"],
)
def get_config(_auth: None = Depends(require_auth)) -> ConfigInfo:
    _, ws = _engine_and_ws()
    yaml_path = ws.config_dir / "company.yaml"
    profile = _read_company_block(yaml_path)
    gmail_ok = gmail_oauth.is_connected(ws)
    drive_ok = gmail_oauth.drive_connected(ws)
    return ConfigInfo(
        mode=_read_mode_from_yaml(yaml_path),
        gmail_connected=gmail_ok,
        drive_connected=drive_ok,
        google_connected=gmail_ok and drive_ok,
        company_configured=bool(profile.name) and bool(profile.one_liner),
    )


@app.post(
    "/config/mode",
    response_model=CommandResult,
    summary="Flip company.yaml `mode:` (fixture | dry_run | production)",
    tags=["onboarding"],
)
def set_mode(
    body: SetModeBody, _auth: None = Depends(require_auth),
) -> CommandResult:
    _, ws = _engine_and_ws()
    try:
        _write_mode_to_yaml(ws.config_dir / "company.yaml", body.mode)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001 - report cleanly to the frontend
        raise HTTPException(
            400,
            detail={
                "error": "failed to update company.yaml",
                "stdout": "",
                "stderr": str(exc),
                "returncode": 1,
            },
        )
    return CommandResult(
        ok=True, returncode=0,
        stdout=f"company.yaml mode set to {body.mode}\n",
        stderr="",
    )


@app.get(
    "/config/company",
    response_model=CompanyProfile,
    summary="Read the company profile (Step 1 form values)",
    tags=["onboarding"],
)
def get_company(_auth: None = Depends(require_auth)) -> CompanyProfile:
    """Returns the flat company profile. Never 404s on missing file --
    a fresh workspace gets the empty shape so the form's controlled
    inputs always have a defined value."""
    _, ws = _engine_and_ws()
    return _read_company_block(ws.config_dir / "company.yaml")


@app.put(
    "/config/company",
    response_model=CommandResult,
    summary="Write the company profile to company.yaml",
    tags=["onboarding"],
)
def put_company(
    body: CompanyProfile, _auth: None = Depends(require_auth),
) -> CommandResult:
    _, ws = _engine_and_ws()
    yaml_path = ws.config_dir / "company.yaml"
    try:
        _write_company_block(yaml_path, body)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001 - report cleanly to the frontend
        raise HTTPException(
            400,
            detail={
                "error": "failed to update company.yaml",
                "stdout": "",
                "stderr": str(exc),
                "returncode": 1,
            },
        )
    return CommandResult(
        ok=True, returncode=0,
        stdout=f"company.yaml `company:` block updated ({body.name or 'unnamed'})\n",
        stderr="",
    )


# ---------- deck-first onboarding extraction ----------
#
# Build Session 15. The wizard's first onboarding step uploads a
# pitch deck; this endpoint parses it, asks the LLM to draft a
# CompanyProfile + per-field evidence, and returns the result. It
# DOES NOT WRITE to company.yaml -- the operator reviews + edits in
# the form, then explicitly calls PUT /config/company to save.

# Cap on the multipart upload. Pitch decks are normally <20MB but
# some image-heavy ones get to 50MB; round to 50MB as the safety
# valve. Files larger than this should be re-exported with
# compressed images.
_MAX_DECK_UPLOAD_BYTES = 50 * 1024 * 1024


def _build_extraction_response(
    *, filename: str, extracted_text, llm_output, llm_warnings: list[str],
) -> "ExtractionResponse":
    """Collapse the LLM's flat extracted_fields list into a draft
    CompanyProfile dict + the audit-trail fields the frontend reads.

    Done as a pure function so the test that doesn't go through HTTP
    can construct the same response without touching the endpoint.
    """
    from schemas.deck_extraction import (
        ExtractionResult,
        NEEDS_REVIEW_THRESHOLD,
        REQUIRED_FIELDS,
    )

    # Start from a default-empty CompanyProfile so the response shape
    # always has every field even if the deck only filled three.
    profile = CompanyProfile().model_dump()
    valid_keys = set(profile.keys())
    extracted_fields = []
    needs_review = set()
    for ef in llm_output.extracted_fields:
        if ef.field not in valid_keys:
            # LLM hallucinated a field name -- skip silently rather
            # than 422'ing the whole response. The evidence-bearing
            # ExtractedField never makes it to the form.
            continue
        # Type-coerce list fields whose CompanyProfile counterpart
        # expects list[str]. The LLM can return a string by mistake;
        # normalize so the frontend doesn't have to.
        if isinstance(profile[ef.field], list) and isinstance(ef.value, str):
            ef = ef.model_copy(update={"value": [ef.value]})
        profile[ef.field] = ef.value
        extracted_fields.append(ef)
        if ef.confidence < NEEDS_REVIEW_THRESHOLD:
            needs_review.add(ef.field)

    missing = []
    for req in REQUIRED_FIELDS:
        val = profile.get(req)
        if val is None or val == "" or val == []:
            missing.append(req)
        # A required field that WAS extracted but with low confidence
        # also goes into needs_review (already added above for
        # confidence; this block just covers the missing case).
    text_preview = extracted_text.text[:1000]

    # ExtractedField (schemas/) and ExtractedFieldOut (web/api.py)
    # are byte-identical in fields but distinct Pydantic models;
    # Pydantic v2 doesn't auto-coerce between them. Round-trip via
    # model_dump so the response_model is happy.
    return ExtractionResponse(
        profile=profile,
        extracted_fields=[ef.model_dump() for ef in extracted_fields],
        missing_required_fields=missing,
        needs_review_fields=sorted(needs_review),
        warnings=extracted_text.warnings + llm_warnings,
        source_filename=filename,
        text_preview=text_preview,
    )


class ExtractedFieldOut(BaseModel):
    """Response copy of schemas.deck_extraction.ExtractedField --
    duplicated here so the FastAPI response_model is self-contained
    (the schemas module sits below the web layer)."""
    field: str
    value: str | int | list[str] | None = None
    confidence: float = 0.0
    evidence: str = ""
    source: str = ""


class ExtractionResponse(BaseModel):
    """HTTP response for POST /config/company/extract-from-deck.

    `profile` is a draft -- the frontend pre-fills the form with
    it, the operator reviews + edits, then PUT /config/company saves.
    The endpoint NEVER writes to company.yaml on its own.
    """
    profile: CompanyProfile
    extracted_fields: list[ExtractedFieldOut] = Field(default_factory=list)
    missing_required_fields: list[str] = Field(default_factory=list)
    needs_review_fields: list[str] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)
    # Review item #21: surface a structured `extraction_failed`
    # flag separate from the freeform warnings list so the
    # frontend can render an unmissable banner ("we couldn't
    # auto-extract; please fill the form manually") on LLM
    # failure -- distinct from the milder "low-confidence field
    # X" warnings that share the warnings list.
    extraction_failed: bool = False
    source_filename: str = ""
    text_preview: str = ""


@app.post(
    "/config/company/extract-from-deck",
    response_model=ExtractionResponse,
    summary="Parse a PDF/PPTX deck and return a draft CompanyProfile",
    tags=["onboarding"],
)
async def extract_from_deck(
    file: UploadFile = File(
        ..., description="PDF or PPTX pitch deck (50 MB max)",
    ),
    _auth: None = Depends(require_auth),
) -> ExtractionResponse:
    """Read the uploaded deck, extract text, ask the LLM to draft
    a CompanyProfile, return the result + per-field evidence.

    Does NOT persist the deck file or mutate company.yaml. The
    response is a SETUP ASSISTANT artifact only; the operator
    reviews + saves via PUT /config/company.
    """
    from core.deck_extraction import extract_profile_draft, extract_text
    from core.llm.client import LLMClient

    content = await file.read()
    if not content:
        raise HTTPException(
            400,
            detail={
                "error": "empty file upload",
                "stdout": "", "stderr": "", "returncode": 1,
            },
        )
    if len(content) > _MAX_DECK_UPLOAD_BYTES:
        raise HTTPException(
            413,
            detail={
                "error": (
                    f"deck is {len(content) // (1024 * 1024)} MB; "
                    f"limit is "
                    f"{_MAX_DECK_UPLOAD_BYTES // (1024 * 1024)} MB. "
                    f"Re-export with compressed images or fill the "
                    f"form manually."
                ),
                "stdout": "", "stderr": "", "returncode": 1,
            },
        )

    extracted = extract_text(file.filename or "", content)
    # When extraction yields no usable text, return a clean empty
    # response with warnings rather than blowing up. The operator
    # falls back to manual entry; PUT /config/company still works.
    if not extracted.text.strip():
        empty_response = _build_extraction_response(
            filename=file.filename or "",
            extracted_text=extracted,
            llm_output=__import__(
                "schemas.deck_extraction",
                fromlist=["DeckLLMOutput"],
            ).DeckLLMOutput(extracted_fields=[]),
            llm_warnings=[],
        )
        # Review #21: same banner signal as LLM-failure -- the
        # operator needs to know auto-extract didn't yield
        # anything actionable before they start editing.
        empty_response.extraction_failed = True
        return empty_response

    _, ws = _engine_and_ws()
    llm = LLMClient(workspace=ws)
    stub = _deck_stub_response() if llm.stub else None
    try:
        llm_output = extract_profile_draft(
            llm=llm, deck_text=extracted.text, stub_response=stub,
        )
        extraction_failed = False
    except Exception as exc:  # noqa: BLE001 - surface as response, not 500
        # An LLM failure during extraction shouldn't 500 the whole
        # onboarding flow -- return what we got from the text layer
        # plus a warning so the operator can continue manually.
        # Review #21: also stamp extraction_failed=True so the
        # frontend can render an unmissable "fill the form
        # manually" banner instead of silently showing an empty
        # profile.
        from schemas.deck_extraction import DeckLLMOutput
        llm_output = DeckLLMOutput(
            extracted_fields=[],
            warnings=[f"LLM extraction failed: {exc}"],
        )
        extraction_failed = True

    response = _build_extraction_response(
        filename=file.filename or "",
        extracted_text=extracted,
        llm_output=llm_output,
        llm_warnings=list(llm_output.warnings),
    )
    response.extraction_failed = extraction_failed
    return response


def _deck_stub_response() -> dict:
    """Stub LLM output for offline mode (CI, tests, no API key).

    Returns a small but realistic ExtractedField set -- enough to
    exercise the endpoint's response shaping (low-confidence -> needs
    review, missing required -> missing_required_fields) without
    requiring a live model.
    """
    return {
        "extracted_fields": [
            {
                "field": "name", "value": "Stub Co",
                "confidence": 0.9,
                "evidence": "(stub) cover slide title",
                "source": "slide 1",
            },
            {
                "field": "one_liner",
                "value": "Stub one-liner from a stub deck.",
                "confidence": 0.8,
                "evidence": "(stub) tagline below the title",
                "source": "slide 1",
            },
            {
                "field": "problem",
                "value": "stub problem statement",
                "confidence": 0.5,   # below NEEDS_REVIEW_THRESHOLD
                "evidence": "(stub) problem section",
                "source": "slide 2",
            },
        ],
        "warnings": [
            "stub-mode response (no ANTHROPIC_API_KEY); fill the "
            "form manually after the deploy is configured with a key"
        ],
    }


def _shell_pipeline_stage(
    script: str, *extra_args: str, label: str,
) -> CommandResult:
    """Shared wrapper for /pipeline/* endpoints. Uses a 10-min timeout
    because Stage 6 (LLM-scored axes) and Stage 7 (LLM-drafted emails)
    both fan out per partner and the request must outlast the slowest
    fixture run."""
    res = _run_cli(
        script, "--workspace", _ws_path(), *extra_args, timeout=600,
    )
    if res.returncode != 0:
        raise HTTPException(
            400,
            detail={
                "error": f"{label} failed",
                "stdout": res.stdout,
                "stderr": res.stderr,
                "returncode": res.returncode,
            },
        )
    return CommandResult(
        ok=True, returncode=res.returncode,
        stdout=res.stdout, stderr=res.stderr,
    )


class PipelineStageResult(BaseModel):
    """One stage's outcome in a multi-stage run. The wizard renders
    each row so the operator can see which stages succeeded /
    failed without parsing combined stdout."""
    stage: str
    ok: bool
    returncode: int = 0
    stdout: str = ""
    stderr: str = ""


class PipelineIngestResult(BaseModel):
    """Response from POST /pipeline/ingest. `stages` is in execution
    order. `ok = all stages ok`."""
    ok: bool
    stages: list[PipelineStageResult]


# Review item #8: scripts that turn a fresh workspace into one
# Stage 6 / 7 can score+draft against. Run in order; any non-zero
# return code aborts the rest and surfaces in `stages[].ok`.
_INGEST_STAGES: list[tuple[str, str]] = [
    ("01_aggregate_sources.py", "aggregate_sources"),
    ("02_enrich_funds.py", "enrich_funds"),
    ("03_mine_activity.py", "mine_activity"),
    ("04_mine_partner_signals.py", "mine_partner_signals"),
    ("05_verify_and_quality.py", "verify_and_quality"),
]


@app.post(
    "/pipeline/ingest",
    response_model=PipelineIngestResult,
    summary=(
        "Run Stages 1-5 (sources -> enrich -> activity -> partner "
        "signals -> verify) end-to-end for the wizard"
    ),
    tags=["onboarding"],
)
def pipeline_ingest(
    _auth: None = Depends(require_auth),
) -> PipelineIngestResult:
    """Wizard's "Run pipeline" button. Fresh workspaces had no
    funds / partners / signals before this -- /pipeline/score and
    /pipeline/generate would either fail or produce empty drafts.
    This endpoint walks Stages 1-5 in order so subsequent /score
    + /generate calls have real data to work with.

    Fail-fast: any stage with returncode != 0 aborts the rest and
    is reported as `ok=False` in the response. The frontend
    surfaces which stage failed.

    Idempotent: re-running over an already-populated workspace
    upserts rather than duplicating.
    """
    out: list[PipelineStageResult] = []
    for script, label in _INGEST_STAGES:
        try:
            cli_res = _run_cli(
                script, "--workspace", _ws_path(),
                *_allow_example_domains_args(),
                timeout=600,
            )
        except Exception as exc:  # noqa: BLE001
            out.append(PipelineStageResult(
                stage=label, ok=False, returncode=1,
                stdout="", stderr=f"runner exception: {exc}",
            ))
            return PipelineIngestResult(ok=False, stages=out)
        ok = cli_res.returncode == 0
        out.append(PipelineStageResult(
            stage=label, ok=ok,
            returncode=cli_res.returncode,
            stdout=cli_res.stdout, stderr=cli_res.stderr,
        ))
        if not ok:
            return PipelineIngestResult(ok=False, stages=out)
    return PipelineIngestResult(ok=True, stages=out)


@app.post(
    "/pipeline/aggregate",
    response_model=CommandResult,
    summary="Run Stage 1 (aggregate_sources) for the wizard",
    tags=["onboarding"],
)
def pipeline_aggregate(
    _auth: None = Depends(require_auth),
) -> CommandResult:
    return _shell_pipeline_stage(
        "01_aggregate_sources.py", label="aggregate_sources",
    )


@app.post(
    "/pipeline/enrich",
    response_model=CommandResult,
    summary="Run Stage 2 (enrich_funds) for the wizard",
    tags=["onboarding"],
)
def pipeline_enrich(
    _auth: None = Depends(require_auth),
) -> CommandResult:
    return _shell_pipeline_stage(
        "02_enrich_funds.py", label="enrich_funds",
    )


@app.post(
    "/pipeline/activity",
    response_model=CommandResult,
    summary="Run Stage 3 (mine_activity) for the wizard",
    tags=["onboarding"],
)
def pipeline_activity(
    _auth: None = Depends(require_auth),
) -> CommandResult:
    return _shell_pipeline_stage(
        "03_mine_activity.py", label="mine_activity",
    )


@app.post(
    "/pipeline/partner-signals",
    response_model=CommandResult,
    summary="Run Stage 4 (mine_partner_signals) for the wizard",
    tags=["onboarding"],
)
def pipeline_partner_signals(
    _auth: None = Depends(require_auth),
) -> CommandResult:
    return _shell_pipeline_stage(
        "04_mine_partner_signals.py", label="mine_partner_signals",
    )


@app.post(
    "/pipeline/verify",
    response_model=CommandResult,
    summary="Run Stage 5 (verify_and_quality) for the wizard",
    tags=["onboarding"],
)
def pipeline_verify(
    _auth: None = Depends(require_auth),
) -> CommandResult:
    return _shell_pipeline_stage(
        "05_verify_and_quality.py", label="verify_and_quality",
    )


@app.post(
    "/pipeline/score",
    response_model=CommandResult,
    summary="Run Stage 6 (score_candidates) for the wizard",
    tags=["onboarding"],
)
def pipeline_score(_auth: None = Depends(require_auth)) -> CommandResult:
    return _shell_pipeline_stage(
        "06_score_candidates.py", label="score_candidates",
    )


@app.post(
    "/pipeline/generate",
    response_model=CommandResult,
    summary="Run Stage 7 (generate_emails) for the wizard",
    tags=["onboarding"],
)
def pipeline_generate(_auth: None = Depends(require_auth)) -> CommandResult:
    # Cap at TOP_BEFORE_CALIBRATION_REQUIRED (=10 in scripts/07) so the
    # wizard's run never trips Gate 5.5's calibration refusal. Fixture
    # domains are allowed only when API_ALLOW_EXAMPLE_DOMAINS is set.
    return _shell_pipeline_stage(
        "07_generate_emails.py",
        "--top", "10",
        *_allow_example_domains_args(),
        label="generate_emails",
    )


# ---------- onboarding step 3: investor sources upload ----------

class SourcesUploadResult(BaseModel):
    """Response from POST /pipeline/sources. The frontend renders
    `row_count` as 'Loaded N investors' confirmation; `stdout` is
    optional diagnostic text the wizard can fold-out behind a 'show
    details' affordance."""
    ok: bool
    row_count: int
    stdout: str = ""


# Cap on the multipart upload. Operator-uploaded investor lists are
# typically a few hundred KB; 25 MB is the safety valve for the case
# where someone tries to upload an Excel export they accidentally
# saved as CSV with embedded BLOBs.
_MAX_SOURCES_UPLOAD_BYTES = 25 * 1024 * 1024


def _sanitize_sources_filename(name: str) -> str:
    """Reduce a user-supplied filename to ASCII-alnum + .csv. Defends
    against path-traversal (../) and Windows / Unix reserved chars
    landing in `data/raw/`. Always ends in `.csv` so Stage 1's CSV
    parser picks it up; falls back to a stable default when the
    incoming name has no usable characters."""
    base = pathlib.Path(name or "").name  # strip any directory parts
    stem = "".join(
        ch for ch in pathlib.Path(base).stem
        if ch.isalnum() or ch in ("_", "-")
    )
    if not stem:
        stem = "operator_sources"
    return f"{stem}.csv"


def _xlsx_to_csv_bytes(content: bytes) -> bytes:
    """Convert an uploaded .xlsx workbook to CSV bytes (UTF-8).

    Reads the first worksheet only -- OpenVC and similar investor
    exports use a single sheet; multi-sheet workbooks aren't a
    documented input format. Headers are lowercased on the way out
    so a sheet whose first row is "Name,Domain" lands as
    "name,domain" -- which is what `scripts/01_aggregate_sources.py`
    expects.
    """
    import csv  # noqa: PLC0415
    import io
    from openpyxl import load_workbook  # noqa: PLC0415

    wb = load_workbook(
        io.BytesIO(content), read_only=True, data_only=True,
    )
    ws = wb.active
    if ws is None:
        raise ValueError("xlsx workbook has no active worksheet")
    out = io.StringIO()
    writer = csv.writer(out)
    first_row_written = False
    for row in ws.iter_rows(values_only=True):
        # Normalize cells: None -> "", non-strings -> str().
        cells = ["" if v is None else str(v) for v in row]
        # Drop trailing all-empty rows (Excel commonly pads with
        # blank rows). A row of "" is content (preserving blanks
        # for downstream tools); a row of nothing at all is noise.
        if not any(c.strip() for c in cells):
            continue
        if not first_row_written:
            # Lowercase the header row so Stage 1's case-sensitive
            # `name` / `domain` lookup matches "Name" / "Domain" /
            # "NAME" exports without per-tool aliases.
            cells = [c.lower().strip() for c in cells]
            first_row_written = True
        writer.writerow(cells)
    return out.getvalue().encode("utf-8")


def _count_csv_rows(content: bytes) -> int:
    """Count data rows (header excluded). Used both to validate the
    CSV parses + to populate `row_count` in the response."""
    import csv  # noqa: PLC0415 - stdlib, hot only when this endpoint fires
    import io
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if len(rows) <= 1:
        # An empty file or a file with only a header has zero data
        # rows. The endpoint refuses upstream because Stage 1 can't
        # do anything with it, but counting is still well-defined.
        return 0
    return len(rows) - 1


# Review items #9 + #10: header / row validation for sources upload.
#
# Stage 1's CSV ingester does case-insensitive header aliasing and
# extracts a fund/firm name + a domain from each row. Pre-this-fix
# `/pipeline/sources` only counted total spreadsheet rows and could
# say "Loaded N investors" when Stage 1 would later ingest zero
# (header names didn't match any alias) -- a silent dead-end.
# These constants are the alias whitelist; keep in sync with the
# downstream sync helper (_sync_uploaded_csv_to_global_pool) and
# Stage 1's parser.
_FIRM_HEADER_ALIASES = (
    "name", "firm", "investor", "investor name",
    "fund", "fund name",
)
_DOMAIN_HEADER_ALIASES = (
    "domain", "website", "url", "homepage", "site",
)


def _count_usable_sources_rows(content: bytes) -> tuple[int, list[str]]:
    """Return (usable_row_count, recognized_headers) for the upload.

    A "usable" row is one that carries at least a firm name -- the
    minimum Stage 1 needs to do anything. Domain is preferred but
    Stage 1 can fall back to enrichment when missing.

    Recognized headers is the intersection of the CSV's header row
    with `_FIRM_HEADER_ALIASES | _DOMAIN_HEADER_ALIASES`; the
    endpoint surfaces it in the error message when no headers
    match, so the operator can rename a column without guessing.
    """
    import csv
    import io
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    raw_headers = reader.fieldnames or []
    normalized = [(h or "").strip().lower() for h in raw_headers]
    accepted = set(_FIRM_HEADER_ALIASES) | set(_DOMAIN_HEADER_ALIASES)
    recognized = [h for h in normalized if h in accepted]
    if not any(h in _FIRM_HEADER_ALIASES for h in normalized):
        # No firm-name column -> Stage 1 cannot ingest a single
        # row. Return 0 + the recognized list (likely empty) so
        # the caller can build a precise error.
        return 0, recognized
    usable = 0
    for row in reader:
        # Lookup is case-insensitive on the alias side. cell values
        # are stripped to drop leading/trailing whitespace.
        lower = {
            (k or "").strip().lower(): (v or "").strip()
            for k, v in row.items() if k is not None
        }
        firm_value = next(
            (lower.get(h) for h in _FIRM_HEADER_ALIASES if lower.get(h)),
            "",
        )
        if firm_value:
            usable += 1
    return usable, recognized


def _link_sources_yaml(yaml_path: pathlib.Path, csv_relative: str,
                       display_name: str) -> None:
    """Prepend an entry to `public_lists` in sources.yaml pointing at
    the uploaded CSV.

    Prepend (not append) so the operator's upload is the first source
    Stage 1 processes -- ahead of fixture seeds. PyYAML drops
    comments; that's acceptable since sources.yaml is rarely
    hand-edited once the wizard is wired up.
    """
    import yaml  # noqa: PLC0415

    data: dict = {}
    if yaml_path.exists():
        text = yaml_path.read_text(encoding="utf-8")
        if text.strip():
            data = yaml.safe_load(text) or {}
    public_lists = data.get("public_lists") or []
    # De-dupe: if a list with the same path is already there, leave it
    # alone (the operator might be re-uploading; idempotent is the
    # right behavior).
    if not any(
        isinstance(item, dict) and item.get("path") == csv_relative
        for item in public_lists
    ):
        public_lists.insert(0, {
            "name": f"Operator upload ({display_name})",
            "path": csv_relative,
            "parser": "csv",
        })
        data["public_lists"] = public_lists
    yaml_path.parent.mkdir(parents=True, exist_ok=True)
    yaml_path.write_text(
        yaml.safe_dump(data, sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )


@app.post(
    "/pipeline/sources",
    response_model=SourcesUploadResult,
    summary="Upload an investor-sources CSV or XLSX; wires it into Stage 1",
    tags=["onboarding"],
)
async def upload_pipeline_sources(
    file: UploadFile = File(
        ..., description="CSV or XLSX of investor sources (25 MB max)",
    ),
    _auth: None = Depends(require_auth),
) -> SourcesUploadResult:
    """Accept the operator's investor-sources file, persist it under
    `clients/<ws>/data/raw/`, and prepend an entry to
    `config/sources.yaml` so the next Stage 1 run picks it up.

    Accepts:
      - `.csv` -- stored as-is.
      - `.xlsx` -- converted to CSV in-memory via openpyxl. Header
        row is lowercased on conversion so OpenVC's `Name`/`Domain`
        columns map to Stage 1's case-sensitive lookup. The
        sources.yaml entry always points at the resulting `.csv` so
        Stage 1's parser doesn't need an xlsx code path.

    Non-destructive: this endpoint never invokes a pipeline stage --
    the wizard's next step is what kicks off Stage 1. Idempotent on
    filename: re-uploading the same name overwrites the file and
    leaves sources.yaml unchanged.
    """
    name = file.filename or ""
    lower = name.lower()
    is_xlsx = lower.endswith(".xlsx")
    is_csv = lower.endswith(".csv")
    if not (is_csv or is_xlsx):
        raise HTTPException(
            400,
            detail={
                "error": (
                    f"sources upload must be a .csv or .xlsx file; "
                    f"got {name!r}"
                ),
                "stdout": "", "stderr": "", "returncode": 1,
            },
        )
    content = await file.read()
    if not content:
        raise HTTPException(
            400,
            detail={
                "error": "empty file upload",
                "stdout": "", "stderr": "", "returncode": 1,
            },
        )
    if len(content) > _MAX_SOURCES_UPLOAD_BYTES:
        raise HTTPException(
            413,
            detail={
                "error": (
                    f"sources upload is "
                    f"{len(content) // (1024 * 1024)} MB; limit is "
                    f"{_MAX_SOURCES_UPLOAD_BYTES // (1024 * 1024)} MB"
                ),
                "stdout": "", "stderr": "", "returncode": 1,
            },
        )

    # XLSX path: convert to CSV bytes in-memory before counting rows.
    # From here on the CSV path is the only persistence shape;
    # sources.yaml always points at `.csv`.
    if is_xlsx:
        try:
            content = _xlsx_to_csv_bytes(content)
        except Exception as exc:  # noqa: BLE001 - openpyxl raises diverse types
            raise HTTPException(
                400,
                detail={
                    "error": (
                        f"could not parse xlsx: {exc}. Re-export as "
                        f".csv or check the file isn't password-"
                        f"protected / corrupt."
                    ),
                    "stdout": "", "stderr": "", "returncode": 1,
                },
            )

    try:
        row_count = _count_csv_rows(content)
    except Exception as exc:  # noqa: BLE001 - csv parse failure
        raise HTTPException(
            400,
            detail={
                "error": f"could not parse CSV: {exc}",
                "stdout": "", "stderr": "", "returncode": 1,
            },
        )
    if row_count <= 0:
        raise HTTPException(
            400,
            detail={
                "error": (
                    "CSV has no data rows (header-only or empty). "
                    "Re-export with at least one investor row."
                ),
                "stdout": "", "stderr": "", "returncode": 1,
            },
        )

    # Review items #9 + #10: validate that the headers Stage 1 will
    # look for are actually present, AND that at least one row
    # has a firm-name value. Without this we silently accept
    # uploads that Stage 1 would then ingest zero rows from.
    try:
        usable_count, recognized = _count_usable_sources_rows(content)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            400,
            detail={
                "error": f"could not parse CSV headers: {exc}",
                "stdout": "", "stderr": "", "returncode": 1,
            },
        )
    if usable_count <= 0:
        # Either no firm-name column at all OR every row had an
        # empty firm cell. Surface the actually-recognized headers
        # (if any) so the operator can rename rather than guess.
        hint = (
            f" recognized headers: {recognized}."
            if recognized else ""
        )
        raise HTTPException(
            400,
            detail={
                "error": (
                    "no recognizable firm/fund-name column. Stage 1 "
                    "looks for one of: "
                    f"{list(_FIRM_HEADER_ALIASES)}." + hint
                ),
                "stdout": "", "stderr": "", "returncode": 1,
            },
        )
    # The frontend renders row_count as 'Loaded N investors'; use
    # the usable count so the count matches what Stage 1 will
    # actually ingest.
    row_count = usable_count

    safe_name = _sanitize_sources_filename(name)
    _, ws = _engine_and_ws()
    raw_dir = ws.path / "data" / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    csv_dst = raw_dir / safe_name
    csv_dst.write_bytes(content)

    csv_relative = f"data/raw/{safe_name}"
    sources_yaml = ws.config_dir / "sources.yaml"
    try:
        _link_sources_yaml(sources_yaml, csv_relative, safe_name)
    except Exception as exc:  # noqa: BLE001 - YAML failure shouldn't drop the upload
        # The CSV is on disk; only the sources.yaml link failed. The
        # operator can hand-edit, but the API should report it rather
        # than silently swallow.
        raise HTTPException(
            500,
            detail={
                "error": f"sources.yaml update failed: {exc}",
                "stdout": (
                    f"CSV saved at {csv_relative} but sources.yaml "
                    f"was not updated; edit it manually to point at "
                    f"the upload."
                ),
                "stderr": str(exc), "returncode": 1,
            },
        )

    # Phase 3: also seed the shared discovery pool. Per the auth
    # spec: every enrichment upserts into investors_global, deduped
    # on email or (firm, partner). The upload is the EARLIEST point
    # we know a firm name + maybe an email, so we seed here; Stage
    # 2/4 enrichment can re-upsert later with richer fields. The
    # sync is best-effort -- a global-pool failure must not block
    # the tenant's CSV landing.
    global_synced = 0
    # Review item #11: per-tenant opt-in check. Skips the sync
    # silently when the tenant hasn't opted in -- the global
    # INVESTORS_GLOBAL_DISABLED env var still acts as the
    # operator-level kill switch beneath it.
    engine, _ = _engine_and_ws()
    with engine.begin() as conn:
        opted_in = _read_discovery_opt_in(conn)
    try:
        if opted_in:
            global_synced = _sync_uploaded_csv_to_global_pool(content)
    except Exception as exc:  # noqa: BLE001 - never block the tenant write
        # Surface in stdout so the operator can audit; don't 5xx
        # the request -- the tenant's CSV is already on disk and
        # sources.yaml is wired.
        global_sync_note = (
            f" (global-pool sync failed: {exc!s}; tenant upload "
            f"succeeded)"
        )
    else:
        # Skip the note entirely on a zero-row sync so the operator
        # can tell "the discovery pool is disabled" / "header-only
        # CSV had nothing to sync" from "synced N rows" without
        # parsing the count. When the tenant opted out, we leave
        # the note blank too -- they get a privacy-by-default
        # silence rather than a "0 rows synced" reveal that the
        # pool exists.
        global_sync_note = (
            f"; synced {global_synced} row(s) into the shared "
            f"discovery pool"
            if global_synced > 0
            else ""
        )

    stdout = (
        f"[sources] uploaded {safe_name} ({row_count} rows) -> "
        f"{csv_relative}; sources.yaml updated{global_sync_note}."
    )
    return SourcesUploadResult(
        ok=True, row_count=row_count, stdout=stdout,
    )


def _sync_uploaded_csv_to_global_pool(csv_bytes: bytes) -> int:
    """Read the just-uploaded CSV and upsert each row into the
    shared `investors_global` pool. Returns the number of rows
    upserted.

    Mapping: the uploaded shape (post-aliasing) carries `name` ==
    firm name and `domain` == fund site host. We seed the global
    pool with one row per (firm, partner=None) -- the discovery
    pool is keyed on firm + partner, so a firm-only row is a
    placeholder Stage 2 / 4 later enriches with partner names +
    emails via a subsequent upsert call (Phase 3 lands the
    mechanism; future PRs wire the per-partner re-sync).

    Disabled when the operator hasn't set GLOBAL_DB_PATH (or in
    tests that haven't enabled the pool) -- the env-default
    `/data/global/global.db` is only meaningful on the Fly volume,
    so a missing path is an explicit signal to skip the sync.
    """
    if os.environ.get("INVESTORS_GLOBAL_DISABLED", "").lower() in (
        "1", "true", "yes", "on",
    ):
        return 0
    import csv as _csv
    import io
    from core.investors_global import (
        InvestorRow,
        get_global_engine,
        upsert_many,
    )
    text = csv_bytes.decode("utf-8-sig", errors="replace")
    reader = _csv.DictReader(io.StringIO(text))
    rows: list[InvestorRow] = []
    for raw_row in reader:
        # The CSV came through with original-case headers; do a
        # case-insensitive lookup for name / firm / investor name.
        lower = {
            (k or "").strip().lower(): (v or "").strip()
            for k, v in raw_row.items() if k is not None
        }
        firm = (
            lower.get("name")
            or lower.get("firm")
            or lower.get("investor")
            or lower.get("investor name")
            or lower.get("fund")
            or lower.get("fund name")
        )
        if not firm:
            continue
        # Partner column is rarely in the upload shape; allow it
        # but default to a placeholder so the firm-only row still
        # gets a stable dedup key.
        partner = (
            lower.get("partner")
            or lower.get("contact")
            or lower.get("partner name")
            or "(unknown)"
        )
        email = (
            lower.get("email")
            or lower.get("partner email")
            or None
        )
        rows.append(InvestorRow(
            firm=firm, partner=partner, email=email,
        ))
    if not rows:
        return 0
    engine = get_global_engine()
    return upsert_many(engine, rows)


# ---------- Phase 4: discovery surface ----------

class DiscoveryMatch(BaseModel):
    """One ranked discovery result. Flat shape for direct frontend
    consumption -- the wizard renders a card per match with fit
    reasons + a Claim button."""
    global_id: int
    firm: str
    partner: str
    email: str | None = None
    stages: list[str] = Field(default_factory=list)
    sectors: list[str] = Field(default_factory=list)
    geographies: list[str] = Field(default_factory=list)
    enriched_fields: dict = Field(default_factory=dict)
    fit_score: int
    fit_reasons: list[str] = Field(default_factory=list)


class DiscoveryMatchesResult(BaseModel):
    matches: list[DiscoveryMatch] = Field(default_factory=list)
    count: int = 0


class DiscoveryClaimBody(BaseModel):
    global_id: int = Field(..., gt=0)


class DiscoveryClaimResult(BaseModel):
    ok: bool
    fund_id: str
    partner_id: str
    global_id: int
    created_fund: bool
    created_partner: bool
    stdout: str = ""


@app.get(
    "/discovery/matches",
    response_model=DiscoveryMatchesResult,
    summary="Ranked investors from the shared pool not yet in your list",
    tags=["discovery"],
)
def discovery_matches(
    limit: int = Query(50, ge=1, le=500),
    _principal: dict | None = Depends(current_principal),
    _auth: None = Depends(require_auth),
) -> DiscoveryMatchesResult:
    """Top-N investors_global rows not already in the tenant's
    `partners` table, ranked by fit to the tenant's company
    profile (sector / stage / geography overlap)."""
    from core.discovery import find_matches
    from core.investors_global import get_global_engine

    engine, ws = _engine_and_ws()
    global_engine = get_global_engine()
    matches = find_matches(
        engine, global_engine, ws.company, limit=limit,
    )
    return DiscoveryMatchesResult(
        matches=[
            DiscoveryMatch(**{
                "global_id": m.global_id,
                "firm": m.firm,
                "partner": m.partner,
                "email": m.email,
                "stages": m.stages,
                "sectors": m.sectors,
                "geographies": m.geographies,
                "enriched_fields": m.enriched_fields,
                "fit_score": m.fit_score,
                "fit_reasons": m.fit_reasons,
            })
            for m in matches
        ],
        count=len(matches),
    )


@app.post(
    "/discovery/claim",
    response_model=DiscoveryClaimResult,
    summary="Copy an investors_global row into your private list",
    tags=["discovery"],
)
def discovery_claim(
    body: DiscoveryClaimBody,
    _principal: dict | None = Depends(current_principal),
    _auth: None = Depends(require_auth),
) -> DiscoveryClaimResult:
    """Idempotent claim: copy the named global investor into the
    tenant's `funds` + `partners` tables, stamping
    `partners.claimed_from_global_id` for audit traceability.

    Re-claiming an already-claimed row is a no-op + still
    returns the same fund_id / partner_id so the frontend can
    deep-link without checking `created_*` flags first."""
    from core.discovery import ClaimError, claim_investor
    from core.investors_global import get_global_engine

    engine, _ = _engine_and_ws()
    global_engine = get_global_engine()
    try:
        result = claim_investor(
            engine, global_engine, body.global_id,
        )
    except ClaimError as exc:
        raise HTTPException(
            404,
            detail={
                "error": str(exc),
                "stdout": "", "stderr": "", "returncode": 1,
            },
        )
    return DiscoveryClaimResult(
        ok=True,
        fund_id=result.fund_id,
        partner_id=result.partner_id,
        global_id=result.global_id,
        created_fund=result.created_fund,
        created_partner=result.created_partner,
        stdout=(
            f"[discovery] claimed global_id={result.global_id} -> "
            f"fund_id={result.fund_id} partner_id={result.partner_id} "
            f"(fund created: {result.created_fund}, "
            f"partner created: {result.created_partner})"
        ),
    )

